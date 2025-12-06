from django.shortcuts import render, redirect , get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.urls import reverse_lazy, reverse
from django.db import models
from django.utils import timezone
from functools import wraps
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import requests
import json
from clientManager.models import Empresa
from .scripts.informe import gen_informe
from .scripts.formularios import procesar_form, revision_form
from loginApp.forms import CrearUsuarioForm, EditarUsuarioForm, CambioClaveAdminForm, ReporteriaForm, FiltrodeFormulariosForm, form_dict, AsignaturaForm, AyudantiaForm
from loginApp.models import Usuario, Asignatura, Ayudantia, Inscripcion, Sede
from solicitudesManager.models import Solicitud

def tutor_required(view_func):
    """
    Decorador que verifica si el usuario es un tutor.
    Redirige a la página de inicio si no es tutor.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_tutor:
            messages.error(request, 'No tienes permisos para acceder a esta sección.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def tenant_user_required(view_func):
    """
    Decorador que verifica que el usuario es del tipo Usuario (tenant) y no AdministradorGlobal.
    Redirige al panel global si es AdministradorGlobal, o al login si no está autenticado.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        
        # Verificar si es AdministradorGlobal
        from clientManager.models import AdministradorGlobal
        if isinstance(request.user, AdministradorGlobal):
            messages.error(request, 'Los administradores globales no pueden acceder a esta sección.')
            return redirect('global_admin:dashboard')
        
        # Verificar que es del tipo Usuario
        if not isinstance(request.user, Usuario):
            messages.error(request, 'Tipo de usuario no válido.')
            return redirect('login')
        
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def index(request):
    # Si está autenticado como AdministradorGlobal y está impersonando, permitir ver el login
    from clientManager.models import AdministradorGlobal
    if request.user.is_authenticated:
        if isinstance(request.user, AdministradorGlobal):
            # Si está impersonando, permitir ver el login del tenant
            # Verificar la sesión directamente para evitar problemas con request.user
            if request.session.get('impersonated_tenant_id'):
                # Continuar mostrando el login del tenant
                pass
            else:
                # Si no está impersonando, redirigir al panel global
                return redirect('global_admin:dashboard')
        else:
            # Si es un usuario normal del tenant, redirigir a home
            return redirect("home")
    
    if request.method == "POST":
        email, clave = request.POST["email"], request.POST["clave"]
        error_email, error_clave = "Email no encontrado", "Contraseña incorrecta"
        errores = {}
        usuario = authenticate(request, email=email, clave=clave)
        if usuario is not None:
            login(request, usuario)
            # Parche para redirecciones del decorador login_required, revisar si hay una mejor solucion
            next_url = request.GET.get('next', '')
            
            # Si hay un parámetro tenant en la sesión o en la URL, mantenerlo en el redirect
            tenant_param = request.GET.get('tenant') or request.session.get('tenant_schema_name')
            if tenant_param:
                if next_url:
                    # Agregar el parámetro tenant a la URL de next
                    separator = '&' if '?' in next_url else '?'
                    next_url = f"{next_url}{separator}tenant={tenant_param}"
                    return HttpResponse(status=302, headers={'Location': next_url})
                else:
                    # Redirigir a home con el parámetro tenant
                    return redirect(f"{reverse('home')}?tenant={tenant_param}")
            
            if next_url:
                return HttpResponse(status=302, headers={'Location': next_url})
            else:
                return redirect("home")
        else:
            try:
                user = Usuario.objects.get(email=email)
                errores["errorClave"] = error_clave
            except Usuario.DoesNotExist:
                errores["errorEmail"] = error_email
            return render(request, "login.html", errores)
    else:
        return render(request, "login.html")
    
@login_required
def home(request):
    usuario = request.user  
    
    # Verificar si es AdministradorGlobal
    from clientManager.models import AdministradorGlobal
    if isinstance(usuario, AdministradorGlobal):
        # Si está impersonando un tenant, permitir navegar el tenant normalmente
        if hasattr(request, 'is_impersonating') and request.is_impersonating:
            # Cuando se está impersonando, mostrar el login del tenant en lugar de redirigir
            # El usuario global puede ver el tenant pero no está autenticado como usuario del tenant
            return redirect('login')  # Redirigir al login del tenant
        else:
            # Si no está impersonando, redirigir al panel global
            return redirect('global_admin:dashboard')
    
    # Verificar que es del tipo Usuario (tenant)
    if not isinstance(usuario, Usuario):
        messages.error(request, 'Tipo de usuario no válido.')
        return redirect('login')
    
    # Redireccionar según el tipo de usuario
    if usuario.is_staff:
        return redirect("usuarios")  # Administradores van al panel de administración
    elif usuario.is_tutor:
        return redirect("tutor_mis_ayudantias")  # Tutores van a sus ayudantías
    else:
        return redirect("estudiante_asignaturas")  # Estudiantes van a ver asignaturas
    
    # Usuarios normales (estudiantes) continúan con la lógica original
    form = tipo_form = None
    empresa = request.tenant  
    casos_empresa = empresa.casos_disponibles if hasattr(empresa, 'casos_disponibles') else []
    if request.method == "POST":
        tipo_form = request.POST.get("tipo_sol")
        if tipo_form in form_dict:
            form_class = form_dict[tipo_form]
            form = form_class(request.POST, request.FILES) if request.FILES else form_class(request.POST)
            if form.is_valid():
                sol = Solicitud(tipo_sol=tipo_form, id_usuario=usuario)
                sol = procesar_form(form, sol)
                lista_sol = Solicitud.objects.filter(estado_sol="Pendiente", tipo_sol=tipo_form)
                if revision_form(sol, lista_sol):
                    sol.save()
                    messages.success(request, "Solicitud guardada con éxito")
                    return redirect("home")
                else:
                    if sol.tipo_sol == "Servicio VPN":
                        usuario, accion = sol.campos_sol["usuario"], sol.campos_sol["accion"]
                        messages.error(request, f"Una solicitud similar existe para el usuario {usuario} con la acción {accion}")
                    else:
                        messages.error(request, "Una solicitud similar existe en curso")
            else:
                messages.error(request, "Por favor corrija los errores en el formulario.")
        else:
            messages.error(request, "Tipo de solicitud no válido.")
            
    form_list = [f() for name, f in form_dict.items() if name in casos_empresa]
    context = {
        "formularios": form_list,
        "current_form": form,  
        "tipo_form": tipo_form,  
    }
    return render(request, "home.html", context)

@login_required
def verSolicitud(request, pk):
    solicitud = get_object_or_404(Solicitud, pk=pk)  
    tipo_solicitud = solicitud.tipo_sol
    form = None  
    if tipo_solicitud in form_dict:
        form = form_dict[tipo_solicitud](initial=solicitud.campos_sol)  
        campos_form = form.fields
        if solicitud.adjunto_sol:
            solicitud.campos_sol["adjunto"] = solicitud.adjunto_sol
        for c in campos_form:
            form.fields[c].widget.attrs['disabled'] = True

    if form: 
        return render(request, "infoSolicitud.html", {"form": form})
    else:
        return redirect("infoSolicitudes")

@login_required
def estadoSolicitudes(request):
    solicitudes = Solicitud.objects.all()
    return render(request, "listaSolicitudes.html", {"solicitudes": solicitudes})

@tenant_user_required
@login_required
def solicitudesUsuario(request):
    try:
        solicitudes = Solicitud.objects.all().filter(id_usuario = request.user.id_usuario)
        return render(request, "listaMisSolicitudes.html", {"solicitudes": solicitudes})
    except Solicitud.DoesNotExist:
        return redirect('home')
     
@login_required
def borrarSolicitud(request, pk):
    solicitud = Solicitud.objects.get(id_sol = pk)
    solicitud.delete()
    return redirect("solicitudes_empresa")

@tenant_user_required
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def usuarios(request):
    solicitudes = Solicitud.objects.filter(id_usuario = request.user.id_usuario)
    usuarios = Usuario.objects.all().order_by('id_usuario')
    datos = {"solicitudes": solicitudes,
             "usuarios": usuarios}
    return render(request, "administrar.html", datos)

@tenant_user_required
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def crearUsuario(request):
    form = CrearUsuarioForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        empresa = Empresa.objects.get(pk=request.tenant.id_empresa)
        usuario = form.save(commit=False)
        usuario.id_empresa = empresa
        usuario.horario_atencion = 0  # Valor por defecto
        usuario.set_password(usuario.password)
        usuario.save()
        return redirect("usuarios")
    solicitudes = Solicitud.objects.filter(id_usuario = request.user.id_usuario)
    datos_solicitudes = {"solicitudes": solicitudes,
                         "formulario": form}
    return render(request, "crearUsuario.html", datos_solicitudes)

@tenant_user_required
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def editarUsuario(request, pk):
    usuario = get_object_or_404(Usuario, id_usuario=pk) 
    if request.method == "POST":
        form = EditarUsuarioForm(request.POST, instance=usuario) 
        if form.is_valid():
            usuario_actualizado = form.save(commit=False)
            # Si is_active está desmarcado, desactivar el usuario
            if not form.cleaned_data.get('is_active', False):
                usuario_actualizado.is_active = False
            usuario_actualizado.save()
            return redirect("usuarios")
        else:
            return render(request, "editarUsuario.html", {"formulario": form})
    else:
        form = EditarUsuarioForm(instance=usuario)
        return render(request, "editarUsuario.html", {"formulario": form})    
   
@tenant_user_required
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def editarClaveAdmin(request, pk):
    usuario = get_object_or_404(Usuario, id_usuario=pk)
    form = CambioClaveAdminForm(request.POST or None, initial={'usuario': usuario})
    if request.method == "POST" and form.is_valid():
        nueva_contraseña = form.cleaned_data['nueva_contraseña']
        usuario.set_password(nueva_contraseña)
        usuario.save()
        update_session_auth_hash(request, usuario)  
        messages.success(request, 'La contraseña ha sido actualizada exitosamente.')
        return render(request, "editarClaveAdmin.html", {"usuario": usuario, "formulario": form})
    elif form.errors:
        messages.error(request, 'Por favor, corrija los errores en el formulario.')

    return render(request, "editarClaveAdmin.html", {"usuario": usuario, "formulario": form})

@tenant_user_required
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def borrarUsuario(request, pk):
    usuario = Usuario.objects.get(id_usuario = pk)
    if request.user.id_usuario == usuario.id_usuario:
        return redirect("usuarios")
    usuario.delete()
    return redirect("usuarios")

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def editarSolicitud(request, pk):
    sol = get_object_or_404(Solicitud, id_sol=pk)
    estados_posibles = ['Pendiente', 'En Proceso', 'Completo']  
    if request.method == "POST":
        estado_sol = request.POST.get('estado_sol')
        tipo_form = sol.tipo_sol 
        if tipo_form in form_dict and estado_sol in estados_posibles:
            form_class = form_dict[tipo_form]
            form = form_class(request.POST, request.FILES, initial=sol.campos_sol)
            if form.is_valid():
                sol.estado_sol = estado_sol
                sol = procesar_form(form, sol)
                lista_sol = Solicitud.objects.filter(estado_sol="Pendiente", tipo_sol=tipo_form).exclude(id_sol=pk)
                if revision_form(sol, lista_sol):
                    sol.save()
                    messages.success(request, "Solicitud actualizada con éxito")
                    return redirect("solicitudes_empresa")
                else:
                    messages.error(request, "Una solicitud similar existe en curso")
            else:
                messages.error(request, "Por favor corrija los errores en el formulario.")
        else:
            messages.error(request, "Tipo de solicitud no válido o estado de solicitud no válido.")
    else:
        if sol.tipo_sol in form_dict:
            tipo_form = form_dict[sol.tipo_sol]
            form = tipo_form(initial=sol.campos_sol)

    context = {
        'formulario': form,
        'tipo_formulario': sol.tipo_sol,
        'estado_sol': sol.estado_sol,
        'estados_posibles': estados_posibles,
    }
    return render(request, "editarSolicitud.html", context)
    
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def casosDeUso(request):
    empresa = Empresa.objects.get(pk=request.tenant.id_empresa)
    if request.method == "POST":
        data = request.POST.getlist("lista_formularios")
        empresa.casos_disponibles = data
        empresa.save()
        return redirect("casosDeUso")
    casos_empresa = empresa.casos_disponibles
    form = FiltrodeFormulariosForm(initial={"lista_formularios": casos_empresa})
    return render(request, "listaCasosdeUso.html", {"formulario": form})
     
@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def reportes(request):
    if request.method == "POST":
        form, response = request.POST, HttpResponse()
        periodo, formato = form["periodo_reportes"], form["formato_reportes"]
        if formato == "pdf":
            response['Content-Disposition'] = "attachment; filename=reporte.pdf"
        else:
            response['Content-Disposition'] = "attachment; filename=reporte.csv"
        gen_informe(periodo, response, formato)
        return response
    
    formulario = ReporteriaForm()
    return render(request, "reporteria.html", {"formulario":formulario})

def logoutProcess(request):
    # Guardar información del tenant antes del logout para usarlo en el template
    tenant_schema = None
    tenant_nombre = None
    if hasattr(request, 'tenant') and request.tenant:
        tenant_schema = request.tenant.schema_name
        tenant_nombre = request.tenant.nombre_empresa
    
    logout(request)
    
    # Limpiar el tenant_schema_name de la sesión para evitar redirecciones automáticas
    # después del logout (solo si existe en la sesión)
    if hasattr(request, 'session') and 'tenant_schema_name' in request.session:
        del request.session['tenant_schema_name']
    
    # Pasar la información del tenant como contexto al template
    return render(request, "logout.html", {
        'tenant_schema': tenant_schema,
        'tenant_nombre': tenant_nombre
    })


def bienvenida(request):
    asignaturas = [
        {
            'nombre': 'Matemáticas',
            'descripcion': 'Clases de álgebra, cálculo, etc.',
            'imagen': 'static/img/matematicas.png'
        },
        {
            'nombre': 'Informática',
            'descripcion': 'Clases de PL/SQL, Java, Python etc.',
            'imagen': 'static/img/informatica.png'
        },
        {
            'nombre': 'Inglés',
            'descripcion': 'Clases de inglés, listening, grammar, etc.',
            'imagen': 'static/img/ingles.png'
        },
        # Agrega más asignaturas según necesites
    ]
    return render(request, 'bienvenida.html', {'asignaturas': asignaturas})

# ========== VISTAS ESPECÍFICAS PARA TUTORES ==========

@tutor_required
def tutor_dashboard(request):
    """
    Dashboard principal para tutores.
    Muestra información relevante para el rol de tutor.
    """
    # Obtener solicitudes asignadas al tutor o relacionadas con su área
    solicitudes = Solicitud.objects.filter(estado_sol="Pendiente").order_by('-created_at')
    
    # Estadísticas básicas
    total_solicitudes = solicitudes.count()
    solicitudes_proceso = Solicitud.objects.filter(estado_sol="En Proceso").count()
    solicitudes_completas = Solicitud.objects.filter(estado_sol="Completo").count()
    
    context = {
        'solicitudes': solicitudes[:10],  # Últimas 10 solicitudes
        'total_solicitudes': total_solicitudes,
        'solicitudes_proceso': solicitudes_proceso,
        'solicitudes_completas': solicitudes_completas,
    }
    return render(request, 'tutor/dashboard.html', context)

@tutor_required
def tutor_solicitudes(request):
    """
    Vista para que los tutores vean y gestionen las solicitudes.
    """
    solicitudes = Solicitud.objects.all().order_by('-created_at')
    return render(request, 'tutor/solicitudes.html', {'solicitudes': solicitudes})

@tutor_required
def tutor_ver_solicitud(request, pk):
    """
    Vista para que los tutores vean los detalles de una solicitud específica.
    """
    solicitud = get_object_or_404(Solicitud, pk=pk)
    tipo_solicitud = solicitud.tipo_sol
    form = None
    
    if tipo_solicitud in form_dict:
        form = form_dict[tipo_solicitud](initial=solicitud.campos_sol)
        campos_form = form.fields
        if solicitud.adjunto_sol:
            solicitud.campos_sol["adjunto"] = solicitud.adjunto_sol
        for c in campos_form:
            form.fields[c].widget.attrs['disabled'] = True

    if form:
        return render(request, "tutor/ver_solicitud.html", {"form": form, "solicitud": solicitud})
    else:
        return redirect("tutor_solicitudes")

@tutor_required
def tutor_actualizar_solicitud(request, pk):
    """
    Vista para que los tutores actualicen el estado de las solicitudes.
    """
    solicitud = get_object_or_404(Solicitud, pk=pk)
    estados_posibles = ['Pendiente', 'En Proceso', 'Completo']
    
    if request.method == "POST":
        estado_sol = request.POST.get('estado_sol')
        observaciones = request.POST.get('observaciones', '')
        
        if estado_sol in estados_posibles:
            solicitud.estado_sol = estado_sol
            if observaciones:
                # Aquí podrías agregar un campo de observaciones al modelo si no existe
                # solicitud.observaciones = observaciones
                pass
            solicitud.save()
            messages.success(request, f"Solicitud actualizada a estado: {estado_sol}")
            return redirect("tutor_ver_solicitud", pk=pk)
        else:
            messages.error(request, "Estado de solicitud no válido.")
    
    context = {
        'solicitud': solicitud,
        'estados_posibles': estados_posibles,
    }
    return render(request, 'tutor/actualizar_solicitud.html', context)

# ========== VISTAS PARA ESTUDIANTES ==========

@login_required
def estudiante_asignaturas(request):
    """
    Vista para que los estudiantes vean las asignaturas disponibles con sus ayudantías.
    Soporta selección de asignatura por parámetro 'asignatura_id'.
    """
    # Obtener asignaturas que tienen ayudantías activas no cursadas
    asignaturas = Asignatura.objects.filter(
        is_active=True,
        ayudantias__is_active=True,
        ayudantias__is_cursada=False,
        ayudantias__fecha__gte=date.today()
    ).distinct().prefetch_related('ayudantias').order_by('nombre')
    
    # Obtener la asignatura seleccionada (si existe)
    asignatura_seleccionada_id = request.GET.get('asignatura_id', None)
    asignatura_seleccionada = None
    if asignatura_seleccionada_id and asignatura_seleccionada_id.strip():
        try:
            # Convertir a entero para la búsqueda
            asignatura_id_int = int(asignatura_seleccionada_id.strip())
            # Buscar directamente en la BD con el filtro correcto
            asignatura_seleccionada = Asignatura.objects.filter(
                is_active=True,
                id_asignatura=asignatura_id_int,
                ayudantias__is_active=True,
                ayudantias__is_cursada=False,
                ayudantias__fecha__gte=date.today()
            ).distinct().prefetch_related('ayudantias').first()
        except (ValueError, TypeError):
            asignatura_seleccionada = None
        except Exception:
            # Capturar cualquier otro error silenciosamente
            asignatura_seleccionada = None
    
    context = {
        'asignaturas': asignaturas,
        'asignatura_seleccionada': asignatura_seleccionada,
        'asignatura_seleccionada_id': asignatura_seleccionada_id,
    }
    return render(request, 'estudiante/asignaturas.html', context)

@login_required
def estudiante_ayudantias_asignatura(request, asignatura_id):
    """
    Vista para mostrar las ayudantías disponibles de una asignatura específica.
    Solo muestra ayudantías no cursadas.
    """
    asignatura = get_object_or_404(Asignatura, id_asignatura=asignatura_id, is_active=True)
    ayudantias = Ayudantia.objects.filter(
        asignatura=asignatura, 
        is_active=True,
        is_cursada=False,  # Solo ayudantías no cursadas
        fecha__gte=date.today()  # Solo ayudantías futuras
    ).order_by('fecha', 'horario')
    
    # Separar ayudantías con y sin cupos
    ayudantias_con_cupos = [a for a in ayudantias if a.cupos_disponibles > 0]
    ayudantias_sin_cupos = [a for a in ayudantias if a.cupos_disponibles <= 0]
    
    context = {
        'asignatura': asignatura,
        'ayudantias_con_cupos': ayudantias_con_cupos,
        'ayudantias_sin_cupos': ayudantias_sin_cupos,
        'total_ayudantias': ayudantias.count(),
    }
    return render(request, 'estudiante/ayudantias_asignatura.html', context)

@login_required
def estudiante_inscribirse(request, ayudantia_id):
    """
    Vista para que un estudiante se inscriba en una ayudantía.
    """
    try:
        ayudantia = Ayudantia.objects.get(id_ayudantia=ayudantia_id, is_active=True, is_cursada=False)
    except Ayudantia.DoesNotExist:
        messages.error(request, 'Esta ayudantía ya no está disponible o fue cursada.')
        return redirect('estudiante_asignaturas')
    
    # Verificar si ya está inscrito
    if Inscripcion.objects.filter(estudiante=request.user, ayudantia=ayudantia).exists():
        messages.warning(request, 'Ya estás inscrito en esta ayudantía.')
        return redirect('estudiante_ayudantias_asignatura', asignatura_id=ayudantia.asignatura.id_asignatura)
    
    # Verificar si hay cupos disponibles
    if ayudantia.cupos_disponibles <= 0:
        messages.error(request, 'No hay cupos disponibles para esta ayudantía.')
        return redirect('estudiante_ayudantias_asignatura', asignatura_id=ayudantia.asignatura.id_asignatura)
    
    # Verificar que la ayudantía sea futura
    if ayudantia.fecha < date.today():
        messages.error(request, 'Esta ayudantía ya pasó.')
        return redirect('estudiante_ayudantias_asignatura', asignatura_id=ayudantia.asignatura.id_asignatura)
    
    # Verificar que la ayudantía no esté cursada
    if ayudantia.is_cursada:
        messages.error(request, 'Esta ayudantía ya fue cursada.')
        return redirect('estudiante_ayudantias_asignatura', asignatura_id=ayudantia.asignatura.id_asignatura)
    
    # Crear la inscripción
    inscripcion = Inscripcion.objects.create(
        estudiante=request.user,
        ayudantia=ayudantia,
        estado='activa'
    )
    
    # Actualizar cupos disponibles
    ayudantia.cupos_disponibles -= 1
    ayudantia.save()
    
    messages.success(request, f'Te has inscrito exitosamente en la ayudantía: {ayudantia.titulo}')
    return redirect('estudiante_mis_ayudantias')

@login_required
def estudiante_mis_ayudantias(request):
    """
    Vista para que el estudiante vea sus ayudantías inscritas.
    Solo muestra ayudantías no cursadas.
    """
    inscripciones = Inscripcion.objects.filter(
        estudiante=request.user, 
        estado='activa',
        ayudantia__is_cursada=False  # Solo ayudantías no cursadas
    ).select_related('ayudantia', 'ayudantia__asignatura', 'ayudantia__tutor')
    
    context = {
        'inscripciones': inscripciones,
    }
    return render(request, 'estudiante/mis_ayudantias.html', context)

@login_required
def estudiante_cancelar_inscripcion(request, inscripcion_id):
    """
    Vista para que el estudiante cancele su inscripción.
    """
    try:
        inscripcion = Inscripcion.objects.get(id_inscripcion=inscripcion_id, estudiante=request.user)
    except Inscripcion.DoesNotExist:
        messages.error(request, 'Esta inscripción ya no existe.')
        return redirect('estudiante_mis_ayudantias')
    
    if inscripcion.estado != 'activa':
        messages.error(request, 'No puedes cancelar una inscripción que no está activa.')
        return redirect('estudiante_mis_ayudantias')
    
    # Verificar que la ayudantía aún existe
    try:
        ayudantia = inscripcion.ayudantia
        # Devolver cupo a la ayudantía solo si aún existe
        ayudantia.cupos_disponibles += 1
        ayudantia.save()
    except Ayudantia.DoesNotExist:
        # La ayudantía fue eliminada, no hay que devolver cupos
        pass
    
    # Eliminar la inscripción completamente para permitir reinscripción
    inscripcion.delete()
    
    messages.success(request, 'Has cancelado tu inscripción exitosamente.')
    return redirect('estudiante_mis_ayudantias')

# ========== VISTAS PARA TUTORES (ACTUALIZADAS) ==========

@tutor_required
def tutor_mis_ayudantias(request):
    """
    Vista para que los tutores vean sus ayudantías.
    Solo muestra ayudantías no cursadas.
    """
    ayudantias = Ayudantia.objects.filter(
        tutor=request.user, 
        is_active=True,
        is_cursada=False,  # Solo ayudantías no cursadas
        fecha__gte=date.today()  # Solo ayudantías futuras
    ).prefetch_related('inscripciones').order_by('fecha', 'horario')
    
    # Calcular inscripciones activas para cada ayudantía
    for ayudantia in ayudantias:
        ayudantia.inscripciones_activas = ayudantia.inscripciones.filter(estado='activa').count()
    
    context = {
        'ayudantias': ayudantias,
    }
    return render(request, 'tutor/mis_ayudantias.html', context)

@tutor_required
def tutor_detalle_ayudantia(request, ayudantia_id):
    """
    Vista para que el tutor vea el detalle de una ayudantía específica.
    Solo muestra ayudantías no cursadas.
    """
    try:
        ayudantia = Ayudantia.objects.get(
            id_ayudantia=ayudantia_id, 
            tutor=request.user, 
            is_active=True,
            is_cursada=False  # Solo ayudantías no cursadas
        )
    except Ayudantia.DoesNotExist:
        messages.error(request, 'Esta ayudantía ya no está disponible o ya fue cursada.')
        return redirect('tutor_mis_ayudantias')
    
    inscripciones = Inscripcion.objects.filter(ayudantia=ayudantia, estado='activa').select_related('estudiante')
    
    # Calcular inscripciones activas
    ayudantia.inscripciones_activas = inscripciones.count()
    
    context = {
        'ayudantia': ayudantia,
        'inscripciones': inscripciones,
    }
    return render(request, 'tutor/detalle_ayudantia.html', context)

@tutor_required
def tutor_registrar_asistencia(request, ayudantia_id):
    """
    Vista para que el tutor registre la asistencia de los estudiantes.
    """
    try:
        ayudantia = Ayudantia.objects.get(
            id_ayudantia=ayudantia_id,
            tutor=request.user,
            is_active=True,
            is_cursada=False  # Solo ayudantías no cursadas
        )
    except Ayudantia.DoesNotExist:
        messages.error(request, 'Esta ayudantía ya no está disponible o fue cursada.')
        return redirect('tutor_mis_ayudantias')
    
    inscripciones = Inscripcion.objects.filter(ayudantia=ayudantia, estado='activa').select_related('estudiante')
    
    if request.method == 'POST':
        # Obtener las asistencias del formulario
        asistencias = request.POST.getlist('asistencia')
        asistio_ids = [int(aid) for aid in asistencias if aid.isdigit()]
        
        # Actualizar las asistencias
        for inscripcion in inscripciones:
            if inscripcion.id_inscripcion in asistio_ids:
                inscripcion.asistio = True
            else:
                inscripcion.asistio = False
            inscripcion.save()
        
        messages.success(request, 'Asistencia registrada exitosamente. Ahora puedes marcar la ayudantía como cursada.')
        return redirect('tutor_detalle_ayudantia', ayudantia_id=ayudantia_id)
    
    context = {
        'ayudantia': ayudantia,
        'inscripciones': inscripciones,
    }
    return render(request, 'tutor/registrar_asistencia.html', context)

@tutor_required
def tutor_marcar_cursada(request, ayudantia_id):
    """
    Vista para que el tutor marque la ayudantía como cursada después de registrar la asistencia.
    """
    try:
        ayudantia = Ayudantia.objects.get(
            id_ayudantia=ayudantia_id,
            tutor=request.user,
            is_active=True,
            is_cursada=False  # Solo ayudantías no cursadas
        )
    except Ayudantia.DoesNotExist:
        messages.error(request, 'Esta ayudantía ya no está disponible o fue cursada.')
        return redirect('tutor_mis_ayudantias')
    
    if request.method == 'POST':
        # Verificar que se haya registrado asistencia al menos para un estudiante
        inscripciones = Inscripcion.objects.filter(ayudantia=ayudantia, estado='activa')
        if not inscripciones.exists():
            messages.error(request, 'No hay estudiantes inscritos en esta ayudantía.')
            return redirect('tutor_detalle_ayudantia', ayudantia_id=ayudantia_id)
        
        # Marcar la ayudantía como cursada
        ayudantia.is_cursada = True
        ayudantia.fecha_cursada = timezone.now()
        ayudantia.save()
        
        messages.success(request, f'Ayudantía "{ayudantia.titulo}" marcada como cursada exitosamente.')
        return redirect('tutor_mis_ayudantias')
    
    # Si es GET, verificar que haya asistencia registrada
    inscripciones = Inscripcion.objects.filter(ayudantia=ayudantia, estado='activa')
    # Verificar si al menos hay una inscripción con asistencia registrada (True o False)
    tiene_asistencia_registrada = inscripciones.filter(asistio__isnull=False).exists()
    
    # Calcular estadísticas de asistencia
    total_inscritos = inscripciones.count()
    total_asistieron = inscripciones.filter(asistio=True).count()
    total_no_asistieron = inscripciones.filter(asistio=False).count()
    
    context = {
        'ayudantia': ayudantia,
        'inscripciones': inscripciones,
        'tiene_asistencia_registrada': tiene_asistencia_registrada,
        'total_inscritos': total_inscritos,
        'total_asistieron': total_asistieron,
        'total_no_asistieron': total_no_asistieron,
    }
    return render(request, 'tutor/marcar_cursada.html', context)

@tutor_required
def tutor_logs(request):
    """
    Vista para que el tutor vea los logs de sus ayudantías cursadas.
    """
    # Obtener filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar ayudantías cursadas del tutor
    ayudantias_cursadas = Ayudantia.objects.filter(
        tutor=request.user,
        is_cursada=True
    ).select_related('asignatura').prefetch_related('inscripciones').order_by('-fecha_cursada')
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__gte=fecha_inicio_obj)
        except ValueError:
            messages.error(request, 'Fecha de inicio inválida.')
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__lte=fecha_fin_obj)
        except ValueError:
            messages.error(request, 'Fecha de fin inválida.')
    
    # Calcular estadísticas para cada ayudantía
    for ayudantia in ayudantias_cursadas:
        inscripciones = ayudantia.inscripciones.filter(estado='activa')
        ayudantia.total_inscritos = inscripciones.count()
        ayudantia.total_asistieron = inscripciones.filter(asistio=True).count()
        ayudantia.total_no_asistieron = inscripciones.filter(asistio=False).count()
    
    context = {
        'ayudantias_cursadas': ayudantias_cursadas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    return render(request, 'tutor/logs.html', context)

@tutor_required
def tutor_exportar_logs_excel(request):
    """
    Vista para exportar los logs del tutor a Excel.
    """
    # Obtener filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar ayudantías cursadas del tutor
    ayudantias_cursadas = Ayudantia.objects.filter(
        tutor=request.user,
        is_cursada=True
    ).select_related('asignatura').prefetch_related('inscripciones').order_by('-fecha_cursada')
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    # Crear el libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Ayudantías"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Título', 'Asignatura', 'Fecha Ayudantía', 'Horario', 'Sala',
        'Fecha Cursada', 'Total Inscritos', 'Total Asistieron', 'Total No Asistieron',
        'Estudiante', 'Email', 'Asistió'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Datos
    row_num = 2
    for ayudantia in ayudantias_cursadas:
        inscripciones = ayudantia.inscripciones.filter(estado='activa')
        total_inscritos = inscripciones.count()
        total_asistieron = inscripciones.filter(asistio=True).count()
        total_no_asistieron = inscripciones.filter(asistio=False).count()
        
        if inscripciones.exists():
            # Una fila por cada inscripción
            for inscripcion in inscripciones:
                ws.cell(row=row_num, column=1).value = ayudantia.id_ayudantia
                ws.cell(row=row_num, column=2).value = ayudantia.titulo
                ws.cell(row=row_num, column=3).value = ayudantia.asignatura.nombre
                ws.cell(row=row_num, column=4).value = ayudantia.fecha.strftime('%d/%m/%Y')
                ws.cell(row=row_num, column=5).value = ayudantia.horario.strftime('%H:%M')
                ws.cell(row=row_num, column=6).value = ayudantia.sala
                ws.cell(row=row_num, column=7).value = ayudantia.fecha_cursada.strftime('%d/%m/%Y %H:%M') if ayudantia.fecha_cursada else ''
                ws.cell(row=row_num, column=8).value = total_inscritos
                ws.cell(row=row_num, column=9).value = total_asistieron
                ws.cell(row=row_num, column=10).value = total_no_asistieron
                ws.cell(row=row_num, column=11).value = inscripcion.estudiante.nombre_usuario
                ws.cell(row=row_num, column=12).value = inscripcion.estudiante.email
                ws.cell(row=row_num, column=13).value = 'Sí' if inscripcion.asistio else 'No'
                row_num += 1
        else:
            # Si no hay inscripciones, mostrar solo la información de la ayudantía
            ws.cell(row=row_num, column=1).value = ayudantia.id_ayudantia
            ws.cell(row=row_num, column=2).value = ayudantia.titulo
            ws.cell(row=row_num, column=3).value = ayudantia.asignatura.nombre
            ws.cell(row=row_num, column=4).value = ayudantia.fecha.strftime('%d/%m/%Y')
            ws.cell(row=row_num, column=5).value = ayudantia.horario.strftime('%H:%M')
            ws.cell(row=row_num, column=6).value = ayudantia.sala
            ws.cell(row=row_num, column=7).value = ayudantia.fecha_cursada.strftime('%d/%m/%Y %H:%M') if ayudantia.fecha_cursada else ''
            ws.cell(row=row_num, column=8).value = 0
            ws.cell(row=row_num, column=9).value = 0
            ws.cell(row=row_num, column=10).value = 0
            ws.cell(row=row_num, column=11).value = 'Sin inscripciones'
            ws.cell(row=row_num, column=12).value = ''
            ws.cell(row=row_num, column=13).value = ''
            row_num += 1
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'logs_ayudantias_tutor_{request.user.id_usuario}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ========== VISTAS PARA ADMINISTRADORES ==========

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_asignaturas(request):
    """
    Vista para que los administradores gestionen las asignaturas.
    Soporta selección de asignatura por parámetro 'asignatura_id'.
    """
    asignaturas = Asignatura.objects.all().prefetch_related('ayudantias').order_by('nombre')
    
    # Calcular si cada asignatura tiene ayudantías
    for asignatura in asignaturas:
        # Verificar si tiene ayudantías activas (no cursadas)
        asignatura.tiene_ayudantias = asignatura.ayudantias.filter(is_active=True, is_cursada=False).exists()
        # Verificar si tiene ayudantías cursadas (para mostrar información adicional)
        asignatura.tiene_ayudantias_cursadas = asignatura.ayudantias.filter(is_cursada=True).exists()
        asignatura.total_ayudantias_cursadas = asignatura.ayudantias.filter(is_cursada=True).count()
    
    # Calcular el número de asignaturas activas (con ayudantías)
    asignaturas_activas = sum(1 for asignatura in asignaturas if asignatura.tiene_ayudantias)
    
    # Obtener la asignatura seleccionada (si existe)
    asignatura_seleccionada_id = request.GET.get('asignatura_id', None)
    asignatura_seleccionada = None
    if asignatura_seleccionada_id:
        try:
            asignatura_seleccionada = Asignatura.objects.get(id_asignatura=asignatura_seleccionada_id)
            # Calcular si tiene ayudantías activas (no cursadas)
            asignatura_seleccionada.tiene_ayudantias = asignatura_seleccionada.ayudantias.filter(is_active=True, is_cursada=False).exists()
            # Verificar si tiene ayudantías cursadas
            asignatura_seleccionada.tiene_ayudantias_cursadas = asignatura_seleccionada.ayudantias.filter(is_cursada=True).exists()
            asignatura_seleccionada.total_ayudantias_cursadas = asignatura_seleccionada.ayudantias.filter(is_cursada=True).count()
            # Total de ayudantías para el template
            asignatura_seleccionada.total_ayudantias = asignatura_seleccionada.ayudantias.count()
        except Asignatura.DoesNotExist:
            asignatura_seleccionada = None
    
    form = AsignaturaForm(request.POST or None)
    
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, 'Asignatura creada exitosamente.')
        return redirect('admin_asignaturas')
    
    context = {
        'asignaturas': asignaturas,
        'asignaturas_activas': asignaturas_activas,
        'form': form,
        'asignatura_seleccionada': asignatura_seleccionada,
        'asignatura_seleccionada_id': asignatura_seleccionada_id,
    }
    return render(request, 'admin/asignaturas.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_editar_asignatura(request, asignatura_id):
    """
    Vista para que los administradores editen asignaturas.
    """
    asignatura = get_object_or_404(Asignatura, id_asignatura=asignatura_id)
    
    if request.method == "POST":
        form = AsignaturaForm(request.POST, instance=asignatura)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura actualizada exitosamente.')
            return redirect('admin_asignaturas')
    else:
        form = AsignaturaForm(instance=asignatura)
    
    context = {
        'form': form,
        'asignatura': asignatura,
    }
    return render(request, 'admin/editar_asignatura.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_eliminar_asignatura(request, asignatura_id):
    """
    Vista para que los administradores eliminen asignaturas.
    Permite eliminar asignaturas incluso si tienen ayudantías cursadas.
    Al eliminar la asignatura, se eliminarán automáticamente todas las ayudantías
    relacionadas (activas y cursadas) y sus inscripciones debido a CASCADE.
    """
    try:
        asignatura = get_object_or_404(Asignatura, id_asignatura=asignatura_id)
        nombre_asignatura = asignatura.nombre
        
        # Contar ayudantías para informar al usuario
        ayudantias_activas = asignatura.ayudantias.filter(is_active=True, is_cursada=False).count()
        ayudantias_cursadas = asignatura.ayudantias.filter(is_cursada=True).count()
        ayudantias_inactivas = asignatura.ayudantias.filter(is_active=False, is_cursada=False).count()
        total_ayudantias = asignatura.ayudantias.count()
        
        # Contar inscripciones totales que se eliminarán
        total_inscripciones = Inscripcion.objects.filter(ayudantia__asignatura=asignatura).count()
        
        # Eliminar la asignatura (esto eliminará automáticamente todas las ayudantías
        # y sus inscripciones debido a on_delete=CASCADE en los modelos)
        asignatura.delete()
        
        # Mensaje de éxito con información sobre lo que se eliminó
        mensaje = f'Asignatura "{nombre_asignatura}" eliminada exitosamente.'
        if total_ayudantias > 0:
            mensaje += f' Se eliminaron {total_ayudantias} ayudantía(s) relacionada(s)'
            if ayudantias_cursadas > 0:
                mensaje += f' ({ayudantias_cursadas} cursada(s) en el historial)'
            if total_inscripciones > 0:
                mensaje += f' y {total_inscripciones} inscripción(es) asociada(s).'
            else:
                mensaje += '.'
        messages.success(request, mensaje)
        
    except Asignatura.DoesNotExist:
        messages.error(request, 'La asignatura no existe.')
    except Exception as e:
        # Capturar cualquier otro error y mostrar un mensaje genérico
        messages.error(request, f'Error al eliminar la asignatura: {str(e)}')
        # Log del error para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error al eliminar asignatura {asignatura_id}: {str(e)}', exc_info=True)
    
    return redirect('admin_asignaturas')

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_crear_ayudantia(request):
    """
    Vista para que los administradores creen ayudantías.
    """
    form = AyudantiaForm(request.POST or None)
    
    # Filtrar solo tutores para el campo tutor
    form.fields['tutor'].queryset = Usuario.objects.filter(is_tutor=True)
    
    if request.method == "POST" and form.is_valid():
        ayudantia = form.save()
        messages.success(request, 'Ayudantía creada exitosamente.')
        return redirect('admin_ayudantias')
    
    context = {
        'form': form,
    }
    return render(request, 'admin/crear_ayudantia.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_ayudantias(request):
    """
    Vista para que los administradores vean todas las ayudantías.
    Solo muestra ayudantías no cursadas.
    """
    ayudantias = Ayudantia.objects.filter(
        is_cursada=False  # Solo ayudantías no cursadas
    ).select_related('tutor', 'asignatura').prefetch_related('inscripciones').order_by('-created_at')
    
    # Calcular inscripciones activas para cada ayudantía
    for ayudantia in ayudantias:
        ayudantia.inscripciones_activas = ayudantia.inscripciones.filter(estado='activa').count()
    
    context = {
        'ayudantias': ayudantias,
    }
    return render(request, 'admin/ayudantias.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_editar_ayudantia(request, ayudantia_id):
    """
    Vista para que los administradores editen ayudantías.
    """
    ayudantia = get_object_or_404(Ayudantia, id_ayudantia=ayudantia_id)
    
    if request.method == "POST":
        form = AyudantiaForm(request.POST, instance=ayudantia)
        # Filtrar solo tutores para el campo tutor
        form.fields['tutor'].queryset = Usuario.objects.filter(is_tutor=True)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Ayudantía actualizada exitosamente.')
            return redirect('admin_ayudantias')
    else:
        form = AyudantiaForm(instance=ayudantia)
        # Filtrar solo tutores para el campo tutor
        form.fields['tutor'].queryset = Usuario.objects.filter(is_tutor=True)
    
    context = {
        'form': form,
        'ayudantia': ayudantia,
    }
    return render(request, 'admin/editar_ayudantia.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_eliminar_ayudantia(request, ayudantia_id):
    """
    Vista para que los administradores eliminen ayudantías.
    """
    try:
        ayudantia = get_object_or_404(Ayudantia, id_ayudantia=ayudantia_id)
        
        # Verificar que la ayudantía no esté cursada
        if ayudantia.is_cursada:
            messages.error(request, 'No se puede eliminar una ayudantía que ya fue cursada. Las ayudantías cursadas se conservan en los logs.')
            return redirect('admin_ayudantias')
        
        titulo_ayudantia = ayudantia.titulo
        
        # Contar inscripciones activas para informar al usuario
        inscripciones_activas = Inscripcion.objects.filter(ayudantia=ayudantia, estado='activa').count()
        
        # Eliminar todas las inscripciones relacionadas primero
        inscripciones_eliminadas = Inscripcion.objects.filter(ayudantia=ayudantia).delete()
        
        # Eliminar la ayudantía
        ayudantia.delete()
        
        # Mensaje de éxito con información sobre inscripciones eliminadas
        if inscripciones_activas > 0:
            messages.success(request, f'Ayudantía "{titulo_ayudantia}" eliminada exitosamente. Se cancelaron {inscripciones_activas} inscripciones de estudiantes.')
        else:
            messages.success(request, f'Ayudantía "{titulo_ayudantia}" eliminada exitosamente.')
        
    except Exception as e:
        messages.error(request, f'Error al eliminar la ayudantía: {str(e)}')
    
    return redirect('admin_ayudantias')

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_logs(request):
    """
    Vista para que el administrador vea los logs de todas las ayudantías cursadas.
    """
    # Obtener filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar todas las ayudantías cursadas
    ayudantias_cursadas = Ayudantia.objects.filter(
        is_cursada=True
    ).select_related('tutor', 'asignatura').prefetch_related('inscripciones').order_by('-fecha_cursada')
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__gte=fecha_inicio_obj)
        except ValueError:
            messages.error(request, 'Fecha de inicio inválida.')
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__lte=fecha_fin_obj)
        except ValueError:
            messages.error(request, 'Fecha de fin inválida.')
    
    # Calcular estadísticas para cada ayudantía
    for ayudantia in ayudantias_cursadas:
        inscripciones = ayudantia.inscripciones.filter(estado='activa')
        ayudantia.total_inscritos = inscripciones.count()
        ayudantia.total_asistieron = inscripciones.filter(asistio=True).count()
        ayudantia.total_no_asistieron = inscripciones.filter(asistio=False).count()
    
    context = {
        'ayudantias_cursadas': ayudantias_cursadas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    return render(request, 'admin/logs.html', context)

@staff_member_required(redirect_field_name=None, login_url=reverse_lazy("home"))
def admin_exportar_logs_excel(request):
    """
    Vista para exportar los logs del admin a Excel.
    """
    # Obtener filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Filtrar todas las ayudantías cursadas
    ayudantias_cursadas = Ayudantia.objects.filter(
        is_cursada=True
    ).select_related('tutor', 'asignatura').prefetch_related('inscripciones').order_by('-fecha_cursada')
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__gte=fecha_inicio_obj)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            ayudantias_cursadas = ayudantias_cursadas.filter(fecha_cursada__date__lte=fecha_fin_obj)
        except ValueError:
            pass
    
    # Crear el libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Ayudantías"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Título', 'Asignatura', 'Tutor', 'Fecha Ayudantía', 'Horario', 'Sala',
        'Fecha Cursada', 'Total Inscritos', 'Total Asistieron', 'Total No Asistieron',
        'Estudiante', 'Email', 'Asistió'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Datos
    row_num = 2
    for ayudantia in ayudantias_cursadas:
        inscripciones = ayudantia.inscripciones.filter(estado='activa')
        total_inscritos = inscripciones.count()
        total_asistieron = inscripciones.filter(asistio=True).count()
        total_no_asistieron = inscripciones.filter(asistio=False).count()
        
        if inscripciones.exists():
            # Una fila por cada inscripción
            for inscripcion in inscripciones:
                ws.cell(row=row_num, column=1).value = ayudantia.id_ayudantia
                ws.cell(row=row_num, column=2).value = ayudantia.titulo
                ws.cell(row=row_num, column=3).value = ayudantia.asignatura.nombre
                ws.cell(row=row_num, column=4).value = ayudantia.tutor.nombre_usuario
                ws.cell(row=row_num, column=5).value = ayudantia.fecha.strftime('%d/%m/%Y')
                ws.cell(row=row_num, column=6).value = ayudantia.horario.strftime('%H:%M')
                ws.cell(row=row_num, column=7).value = ayudantia.sala
                ws.cell(row=row_num, column=8).value = ayudantia.fecha_cursada.strftime('%d/%m/%Y %H:%M') if ayudantia.fecha_cursada else ''
                ws.cell(row=row_num, column=9).value = total_inscritos
                ws.cell(row=row_num, column=10).value = total_asistieron
                ws.cell(row=row_num, column=11).value = total_no_asistieron
                ws.cell(row=row_num, column=12).value = inscripcion.estudiante.nombre_usuario
                ws.cell(row=row_num, column=13).value = inscripcion.estudiante.email
                ws.cell(row=row_num, column=14).value = 'Sí' if inscripcion.asistio else 'No'
                row_num += 1
        else:
            # Si no hay inscripciones, mostrar solo la información de la ayudantía
            ws.cell(row=row_num, column=1).value = ayudantia.id_ayudantia
            ws.cell(row=row_num, column=2).value = ayudantia.titulo
            ws.cell(row=row_num, column=3).value = ayudantia.asignatura.nombre
            ws.cell(row=row_num, column=4).value = ayudantia.tutor.nombre_usuario
            ws.cell(row=row_num, column=5).value = ayudantia.fecha.strftime('%d/%m/%Y')
            ws.cell(row=row_num, column=6).value = ayudantia.horario.strftime('%H:%M')
            ws.cell(row=row_num, column=7).value = ayudantia.sala
            ws.cell(row=row_num, column=8).value = ayudantia.fecha_cursada.strftime('%d/%m/%Y %H:%M') if ayudantia.fecha_cursada else ''
            ws.cell(row=row_num, column=9).value = 0
            ws.cell(row=row_num, column=10).value = 0
            ws.cell(row=row_num, column=11).value = 0
            ws.cell(row=row_num, column=12).value = 'Sin inscripciones'
            ws.cell(row=row_num, column=13).value = ''
            ws.cell(row=row_num, column=14).value = ''
            row_num += 1
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'logs_ayudantias_admin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ========== VISTA DE TEST PARA API DE FERIADOS ==========

@login_required
def test_api_feriados(request):
    """
    Vista de prueba para consumir la API de feriados de Boostr.
    """
    api_url = "https://api.boostr.cl/holidays.json"
    feriados = []
    error = None
    status_api = None
    
    try:
        # Realizar petición a la API
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()  # Lanza excepción si hay error HTTP
        
        # Parsear la respuesta JSON
        data = response.json()
        status_api = data.get('status', 'unknown')
        
        if status_api == 'success' and 'data' in data:
            feriados = data['data']
        else:
            error = f"La API retornó un estado inesperado: {status_api}"
            
    except requests.exceptions.Timeout:
        error = "La petición a la API tardó demasiado tiempo."
    except requests.exceptions.RequestException as e:
        error = f"Error al conectar con la API: {str(e)}"
    except json.JSONDecodeError:
        error = "Error al parsear la respuesta JSON de la API."
    except Exception as e:
        error = f"Error inesperado: {str(e)}"
    
    # Separar feriados por tipo para mejor visualización
    feriados_civiles = [f for f in feriados if f.get('type') == 'Civil']
    feriados_religiosos = [f for f in feriados if f.get('type') == 'Religioso']
    feriados_inalienables = [f for f in feriados if f.get('inalienable') == True]
    
    # Ordenar feriados por fecha
    feriados.sort(key=lambda x: x.get('date', ''))
    
    context = {
        'feriados': feriados,
        'feriados_civiles': feriados_civiles,
        'feriados_religiosos': feriados_religiosos,
        'feriados_inalienables': feriados_inalienables,
        'total_feriados': len(feriados),
        'error': error,
        'status_api': status_api,
        'api_url': api_url,
    }
    
    return render(request, 'test_api_feriados.html', context)

# ========== VISTA DE TEST PARA MAPA DE SEDES ==========

@login_required
def test_mapa_sedes(request):
    """
    Vista para mostrar un mapa con las sedes de la institución usando Leaflet.
    """
    # Obtener todas las sedes activas
    sedes = Sede.objects.filter(is_active=True)
    
    # Crear una lista de diccionarios con los datos de las sedes
    lista_sedes = []
    for sede in sedes:
        lista_sedes.append({
            'id': sede.id_sede,
            'nombre': sede.nombre,
            'lat': sede.latitud,
            'lng': sede.longitud,
            'direccion': sede.direccion
        })
    
    # Convertir a JSON para enviarlo al template
    sedes_json = json.dumps(lista_sedes, ensure_ascii=False)
    
    context = {
        'sedes_json': sedes_json,
        'total_sedes': len(lista_sedes),
        'sedes': sedes,  # También pasamos el queryset por si lo necesitamos
    }
    
    return render(request, 'test_mapa_sedes.html', context)